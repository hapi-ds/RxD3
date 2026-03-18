"""
FMEA Report Service for generating XLSX reports from graph data.

This module implements the FMEAReportService class that traverses the
Mind Graph's failure chains (Requirement ← CAN_OCCUR ← Risk ← LEAD_TO ← Failure)
and generates FMEA XLSX reports using openpyxl templates.

**Validates: Requirements 6.1–6.16**
"""

import io
import logging
from pathlib import Path
from typing import Any, Optional

from neontology import GraphConnection
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Resolve project root (backend/ is one level up from src/)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


class FMEAReportService:
    """
    Service for generating FMEA XLSX reports from Mind Graph data.

    Traverses the graph pattern Requirement ←[CAN_OCCUR]— Risk ←[LEAD_TO]— Failure,
    following failure chains via DFS, and fills XLSX templates with one row per
    unique path to a terminal Failure node.

    **Validates: Requirements 6.1–6.16**
    """

    TEMPLATE_MAP: dict[str, str] = {
        "design": "templates/Template_dFMEA.xlsx",
        "process": "templates/Template_pFMEA.xlsx",
        "iso14971": "templates/Template_iFMEA.xlsx",
        "general": "templates/Template.xlsx",
    }

    # Column mappings for design/process/general templates (29 columns)
    _STANDARD_COLUMNS: dict[str, int] = {
        "requirement": 3,
        "effects": 4,
        "severity": 5,
        "failure_mode": 6,
        "causes": 7,
        "mitigation_plan": 8,
        "occurrence": 9,
        "detection_method": 10,
        "detectability": 11,
        "p1": 25,
        "p2": 26,
        "cumulative_occurrence": 27,
        "cumulative_detectability": 28,
        "acceptable_limit": 29,
    }

    # Column mappings for iso14971 template (25 columns, no detectability)
    _ISO14971_COLUMNS: dict[str, int] = {
        "requirement": 3,
        "effects": 4,
        "severity": 5,
        "failure_mode": 6,
        "causes": 7,
        "mitigation_plan": 8,
        "occurrence": 9,
        "p1": 22,
        "p2": 23,
        "cumulative_occurrence": 24,
        "acceptable_limit": 25,
    }

    # Header row in all templates; data starts on the next row
    _HEADER_ROW = 10

    def generate_report(self, fmea_type: str) -> bytes:
        """Generate an FMEA XLSX report for the given type.

        Args:
            fmea_type: One of "design", "process", "iso14971", "general".

        Returns:
            The XLSX file content as bytes.

        Raises:
            ValueError: If fmea_type is not recognised.
            FileNotFoundError: If the template file is missing.
            RuntimeError: If no FMEA data exists in the graph.
        """
        if fmea_type not in self.TEMPLATE_MAP:
            raise ValueError(
                f"Invalid FMEA type: {fmea_type}. "
                f"Must be one of: {', '.join(sorted(self.TEMPLATE_MAP))}"
            )

        template_rel = self.TEMPLATE_MAP[fmea_type]
        template_path = _PROJECT_ROOT / template_rel
        if not template_path.exists():
            raise FileNotFoundError(f"FMEA template not found: {template_path}")

        rows = self._traverse_failure_chains()
        if not rows:
            raise RuntimeError(
                "No FMEA data available. Load example data or create "
                "Requirements with associated Risks."
            )

        is_iso = fmea_type == "iso14971"
        col_map = self._ISO14971_COLUMNS if is_iso else self._STANDARD_COLUMNS

        wb = load_workbook(template_path)
        ws = wb.active
        data_start_row = self._HEADER_ROW + 1

        for idx, row_data in enumerate(rows):
            row_num = data_start_row + idx
            ws.cell(row=row_num, column=col_map["requirement"], value=row_data["requirement"])
            ws.cell(row=row_num, column=col_map["effects"], value=row_data["effects"])
            ws.cell(row=row_num, column=col_map["severity"], value=row_data["severity"])
            ws.cell(row=row_num, column=col_map["failure_mode"], value=row_data["failure_mode"])
            ws.cell(row=row_num, column=col_map["causes"], value=row_data["causes"])
            ws.cell(row=row_num, column=col_map["mitigation_plan"], value=row_data.get("mitigation_plan"))
            ws.cell(row=row_num, column=col_map["occurrence"], value=row_data.get("occurrence"))
            ws.cell(row=row_num, column=col_map["p1"], value=row_data.get("p1"))
            ws.cell(row=row_num, column=col_map["p2"], value=row_data.get("p2"))
            ws.cell(
                row=row_num,
                column=col_map["cumulative_occurrence"],
                value=row_data["cumulative_occurrence"],
            )
            ws.cell(row=row_num, column=col_map["acceptable_limit"], value=row_data.get("acceptable_limit"))

            if not is_iso:
                ws.cell(row=row_num, column=col_map["detection_method"], value=row_data.get("detection_method"))
                ws.cell(row=row_num, column=col_map["detectability"], value=row_data.get("detectability"))
                ws.cell(
                    row=row_num,
                    column=col_map["cumulative_detectability"],
                    value=row_data["cumulative_detectability"],
                )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _traverse_failure_chains(self) -> list[dict[str, Any]]:
        """Traverse the graph to collect failure chain rows via DFS."""
        gc = GraphConnection()

        # Step 1 & 2: Get all (Requirement, Risk, CanOccur) triples and Optional Mitigation
        req_risk_query = """
        MATCH (risk)-[co:CAN_OCCUR]->(req:Requirement)
        OPTIONAL MATCH (m:Mitigation)-[:MITIGATES]->(risk)
        RETURN req.title AS req_title,
               risk.title AS risk_title,
               risk.description AS effects,
               risk.severity AS severity,
               m.title AS mitigation_plan,
               co.p1 AS p1,
               co.p2 AS p2,
               elementId(risk) AS risk_eid
        """
        req_risk_results = gc.engine.evaluate_query(req_risk_query, {})

        if not req_risk_results or not req_risk_results.records_raw:
            return []

        rows: list[dict[str, Any]] = []

        for record in req_risk_results.records_raw:
            risk_eid = record["risk_eid"]
            base_info: dict[str, Any] = {
                "requirement": record["req_title"],
                "effects": record["effects"] or record["risk_title"],
                "severity": record["severity"],
                "mitigation_plan": record.get("mitigation_plan"),
                "acceptable_limit": None,
                "p1": record.get("p1"),
                "p2": record.get("p2"),
            }

            # Step 3: Find nearest Failure nodes that LEAD_TO this Risk
            failure_to_risk_query = """
            MATCH (f:Failure)-[lt:LEAD_TO]->(risk)
            WHERE elementId(risk) = $risk_eid
            RETURN elementId(f) AS f_eid,
                   f.title AS failure_mode,
                   f.occurrence AS occurrence,
                   f.detectability AS detectability,
                   lt.occurrence_probability AS occ_prob,
                   lt.detectability_probability AS det_prob
            """
            f_results = gc.engine.evaluate_query(
                failure_to_risk_query, {"risk_eid": risk_eid}
            )

            if not f_results or not f_results.records_raw:
                continue

            for f_rec in f_results.records_raw:
                chain_probs: list[dict[str, Optional[float]]] = [
                    {
                        "occ_prob": f_rec.get("occ_prob"),
                        "det_prob": f_rec.get("det_prob"),
                    }
                ]
                
                failure_data = {
                    "failure_mode": f_rec["failure_mode"],
                    "occurrence": f_rec.get("occurrence"),
                    "detectability": f_rec.get("detectability"),
                }
                
                causes_chain: list[str] = []

                self._dfs_failure_chains(
                    gc,
                    f_rec["f_eid"],
                    failure_data,
                    chain_probs,
                    causes_chain,
                    base_info,
                    rows,
                )

        return rows

    def _dfs_failure_chains(
        self,
        gc: GraphConnection,
        failure_eid: str,
        failure_data: dict[str, Any],
        chain_probs: list[dict[str, Optional[float]]],
        causes_chain: list[str],
        base_info: dict[str, Any],
        rows: list[dict[str, Any]],
    ) -> None:
        """Recursively follow LEAD_TO from Failure → Failure via DFS."""
        upstream_query = """
        MATCH (f:Failure)-[lt:LEAD_TO]->(target)
        WHERE elementId(target) = $target_eid
          AND f:Failure
        RETURN elementId(f) AS f_eid,
               f.title AS cause_title,
               lt.occurrence_probability AS occ_prob,
               lt.detectability_probability AS det_prob
        """
        upstream_results = gc.engine.evaluate_query(
            upstream_query, {"target_eid": failure_eid}
        )

        has_upstream = (
            upstream_results
            and upstream_results.records_raw
            and len(upstream_results.records_raw) > 0
        )

        if not has_upstream:
            cum = self._calculate_cumulative_probabilities(chain_probs)
            causes_str = " <- ".join(reversed(causes_chain)) if causes_chain else "None"
            
            rows.append(
                {
                    "requirement": base_info["requirement"],
                    "severity": base_info["severity"],
                    "effects": base_info["effects"],
                    "failure_mode": failure_data["failure_mode"],
                    "causes": causes_str,
                    "occurrence": failure_data.get("occurrence"),
                    "detectability": failure_data.get("detectability"),
                    "detection_method": None,
                    "mitigation_plan": base_info.get("mitigation_plan"),
                    "acceptable_limit": None,
                    "p1": base_info.get("p1"),
                    "p2": base_info.get("p2"),
                    **cum,
                }
            )
        else:
            for u_rec in upstream_results.records_raw:
                new_probs = chain_probs + [
                    {
                        "occ_prob": u_rec.get("occ_prob"),
                        "det_prob": u_rec.get("det_prob"),
                    }
                ]
                new_causes_chain = causes_chain + [u_rec["cause_title"]]
                
                self._dfs_failure_chains(
                    gc,
                    u_rec["f_eid"],
                    failure_data,
                    new_probs,
                    new_causes_chain,
                    base_info,
                    rows,
                )

    @staticmethod
    def _calculate_cumulative_probabilities(
        chain: list[dict[str, Optional[float]]],
    ) -> dict[str, Any]:
        """Multiply occurrence and detectability probabilities along the chain.

        If any value in the chain is None/missing, the cumulative result is "na".

        Args:
            chain: List of dicts with "occ_prob" and "det_prob" keys.

        Returns:
            Dict with "cumulative_occurrence" and "cumulative_detectability" values.

        **Validates: Requirements 6.10, 6.11, 6.14**
        """
        cum_occ: Any = 1.0
        cum_det: Any = 1.0
        occ_missing = False
        det_missing = False

        for link in chain:
            occ = link.get("occ_prob")
            det = link.get("det_prob")

            if occ is None:
                occ_missing = True
            elif not occ_missing:
                cum_occ *= occ

            if det is None:
                det_missing = True
            elif not det_missing:
                cum_det *= det

        return {
            "cumulative_occurrence": "na" if occ_missing else cum_occ,
            "cumulative_detectability": "na" if det_missing else cum_det,
        }
